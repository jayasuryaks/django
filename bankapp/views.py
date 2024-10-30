from django.contrib.auth.hashers import make_password, check_password
from django.shortcuts import render, redirect
from django.contrib import messages
import random
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from .models import BankModel  # Ensure your model is imported


# Registration view
def index(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        account_type = request.POST.get('accountType')
        # address = request.POST.get('address')  # Assuming address is part of your form
        deposit = request.POST.get('deposit')

        if all([username, email, phone, account_type, deposit]):

            # Generate a random password
            deposit=float(deposit)
            random_number = random.randint(1000, 9999)  # Ensures 4 digits
            password = f"{random_number:04d}"
            hashed_password = make_password(password)

            # Generate a unique account number
            base_account_number = "1000210001"
            random_five_digits = f"{random.randint(0, 99999):05d}"
            account_number = base_account_number + random_five_digits
            print(password)
            user = BankModel(
                username=username,
                email=email,
                phone=phone,
                account_type=account_type,
                account_number=account_number,
                password=hashed_password,
                deposit=deposit
            )
            user.save()
            print(user)
            messages.success(request, 'Registration successful! Your account number is: ' + account_number)

            def save_to_excel(name, account_number, email, password):
                """Saves user data to an Excel file."""
                try:
                    workbook = openpyxl.load_workbook('input_data.xlsx')
                    sheet = workbook.active
                except FileNotFoundError:
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.title = 'User Input'

                    # Add headers to the new sheet
                    headers = ['Name', 'Account Number', 'Email', 'Password']
                    sheet.append(headers)

                    # Format headers
                    for cell in sheet[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                except InvalidFileException:
                    print("The file 'input_data.xlsx' is corrupted or invalid.")
                    return

                # Append the new user data to the next row
                sheet.append([name, account_number, email, password])

                # Save the workbook
                workbook.save('input_data.xlsx')
                print("Data saved to input_data.xlsx")

            return redirect('login')  # Redirect to login view

    return render(request, "index.html")


# Login view
def login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        try:
            user = BankModel.objects.get(email=email)
            if check_password(password, user.password):
                request.session['email'] = email
                return redirect('profile')
            else:
                messages.error(request, 'Invalid password.')
        except BankModel.DoesNotExist:
            messages.error(request, 'Email not found.')

    return render(request, 'login.html')


# Profile view
def profile(request):
    if 'email' not in request.session:
        return redirect('login')

    email = request.session['email']
    user = BankModel.objects.get(email=email)


    return render(request, 'profile.html', {'user': user})


# Logout view
def logout(request):
    if 'email' in request.session:
        del request.session['email']
        messages.success(request, 'You have successfully logged out.')

    return redirect('login')  # Redirect to login after logout
